#!/usr/bin/env python3
"""
Vridik / JuliX — eval/evaluador.py
Sprint S5: banco de evaluación de JuliX (GATE de Fase 1).

Qué hace:
  1. Lee eval/banco_casos_vridik.xlsx (20 casos: 12 UGPP + 8 Laboral).
  2. Para cada caso CON patrón oro ya llenado por Ana Luisa (columna
     `respuesta_esperada`, ver eval/guia_abogada.md), genera la respuesta de
     JuliX usando el prompt real de producción (julix/prompts/) y el modelo
     confirmado `claude-sonnet-5` (julix/client.py).
  3. Le pide a un "Claude juez" (mismo cliente, tarea `evaluacion_juez`) que
     califique 0-5 la respuesta de JuliX contra el patrón oro y la
     `norma_clave`, evaluando: precisión normativa, ausencia de alucinación
     (¿cita algo que no está en norma_clave?) y si la cita es correcta.
  4. Persiste cada evaluación en `julix_evals` (eval/sql/julix_evals_schema.sql):
     score, hallucination_flag, costo_usd (generación + juez).
  5. Calcula el % de aprobación de la corrida y aplica el GATE de Fase 1:
     >= 80% de casos aprobados (score >= 4 y sin hallucination_flag).

Un caso se considera "aprobado" si score >= UMBRAL_APROBACION_CASO_SCORE (4)
Y hallucination_flag es False. Una alucinación nunca se compensa con un
score alto en las otras dimensiones (regla heredada del roadmap de S5:
"alucinación = global 1 automático").

Modo por defecto: --dry-run (no llama a Claude, no escribe en BD; solo
valida el banco, cuenta cuántos casos ya tienen patrón oro y muestra el
plan). Usar --commit para la corrida real.

USO:
    python eval/evaluador.py --excel eval/banco_casos_vridik.xlsx
    python eval/evaluador.py --excel eval/banco_casos_vridik.xlsx --commit

NO SE EJECUTA CONTRA ANTHROPIC REAL NI CONTRA UNA BASE DE DATOS REAL EN ESTE
ENTREGABLE.
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import logging
import sys
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

# Repo layout: julix/, eval/ son hermanos en la raíz.
REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from julix import prompts  # noqa: E402
from julix.client import JuliXClient  # noqa: E402
from julix.errors import JuliXError, JuliXInvalidFormatError  # noqa: E402
from julix.service import _claves_citables  # noqa: E402

logger = logging.getLogger("vridik.julix.eval")

UMBRAL_APROBACION_CASO_SCORE = 4  # score mínimo (0-5) para considerar un caso aprobado
GATE_FASE1_PORCENTAJE = 0.80       # Gate de Fase 1: >=80% de casos aprobados
# None, no un sentinel de texto: julix_calls.user_id es UUID con FK real a
# users(id) -- el banco de evaluación no corre como ningún usuario real de
# la app, así que NULL (nullable a propósito, ver julix/ledger.py::ensure_
# julix_calls_table) es lo correcto, no un string inventado (bug real
# encontrado el 15-jul-2026 en la primera corrida real de este script:
# "banco_evaluacion_s5" no es un UUID válido, rompía el INSERT).
USER_ID_BANCO = None

TAREA_POR_AREA = {
    "UGPP": "ugpp_demanda",
    "Laboral": "laboral_consulta",
}

JUEZ_SYSTEM_PROMPT = """Eres el juez de calidad del banco de evaluación de Vridik/JuliX.

Recibes: una pregunta jurídica, la respuesta que dio JuliX, el patrón oro
(la respuesta que el despacho considera correcta) y la norma_clave (la
única fuente normativa que JuliX tenía permitido citar en este ejercicio).

Califica la respuesta de JuliX en estas dimensiones:
  - precision_normativa (0-5): ¿la respuesta identifica correctamente la
    norma y su contenido, comparada con el patrón oro?
  - cita_correcta (true/false): ¿toda cita normativa en la respuesta de
    JuliX corresponde EXACTAMENTE a la norma_clave entregada (no a otra
    norma inventada o distinta)?
  - hallucination_flag (true/false): ¿JuliX afirma algo (una norma, una
    cifra, un plazo) que no está respaldado por la norma_clave entregada?
    Una alucinación es cualquier afirmación normativa sin respaldo en la
    fuente permitida, exista o no en la realidad — en este ejercicio la
    única fuente válida es la norma_clave dada.

    IMPORTANTE (chequeo mecánico, no de criterio): antes de marcar
    hallucination_flag=true por una norma/cita específica, verificá
    literalmente si esa cita (ley, decreto o artículo) aparece dentro del
    texto de "norma_clave" que se te entregó abajo. Si aparece ahí, JuliX
    NO la inventó — fue entregada como fuente autorizada, y comentar sobre
    su vigencia, plausibilidad o fecha NO es alucinación (puede ser
    prudencia legítima). Solo marcá hallucination_flag=true por normas que
    JuliX introdujo y que NO están en el texto de norma_clave.
  - score (0-5): calificación global — si hallucination_flag es true, el
    score NUNCA puede ser mayor a 1, sin excepción.
  - comentario: 1-2 frases explicando la calificación.

Responde ÚNICAMENTE con un objeto JSON con exactamente estas claves:
{"score": <0-5>, "precision_normativa": <0-5>, "cita_correcta": <true|false>, "hallucination_flag": <true|false>, "comentario": "<texto breve>"}

No agregues texto antes ni después del JSON.
"""


@dataclass
class CasoEval:
    id: str
    area: str
    pregunta: str
    respuesta_esperada: str
    norma_clave: str
    dificultad: int


@dataclass
class ResultadoCaso:
    caso_id: str
    area: str
    dificultad: int
    model: str
    score: int
    precision_normativa: int
    cita_correcta: bool
    hallucination_flag: bool
    comentario_juez: str
    respuesta_julix: str
    costo_usd_generacion: float | None
    costo_usd_juez: float | None


@dataclass
class ResumenBanco:
    run_id: str
    total_casos: int
    casos_evaluados: int
    casos_pendientes_patron_oro: int
    casos_aprobados: int
    casos_con_alucinacion: int
    porcentaje_aprobacion: float
    gate_aprobado: bool
    resultados: list[ResultadoCaso] = field(default_factory=list)


def leer_banco(excel_path: Path) -> list[CasoEval]:
    if load_workbook is None:
        raise RuntimeError("Falta la dependencia 'openpyxl' (pip install openpyxl)")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb["Banco de Casos"]

    casos: list[CasoEval] = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        id_, area, pregunta, respuesta_esperada, norma_clave, dificultad = fila[:6]
        if id_ is None:
            continue
        casos.append(
            CasoEval(
                id=str(id_),
                area=str(area),
                pregunta=str(pregunta or ""),
                respuesta_esperada=str(respuesta_esperada or "").strip(),
                norma_clave=str(norma_clave or ""),
                dificultad=int(dificultad) if dificultad else 1,
            )
        )
    return casos


def _hash_prompt(texto: str) -> str:
    return hashlib.sha256(texto.encode("utf-8")).hexdigest()[:16]


async def generar_respuesta_julix(client: JuliXClient, caso: CasoEval) -> tuple[str, str]:
    """Genera la respuesta de JuliX para un caso usando el prompt real de
    producción (julix/prompts/). Retorna (texto_respuesta, tarea)."""
    tarea = TAREA_POR_AREA[caso.area]
    prompt = prompts.load_prompt(tarea)

    # El banco de S5 usa como única fuente permitida la norma_clave del caso
    # (no el corpus completo, que llega en S7-S9) — esto es justamente lo
    # que permite medir alucinación: cualquier cita fuera de norma_clave es
    # una alucinación por definición en este ejercicio.
    user_content = (
        f"Pregunta del caso ({caso.id}, área {caso.area}):\n{caso.pregunta}\n\n"
        f"Fuente normativa permitida para esta respuesta (única fuente válida):\n{caso.norma_clave}"
    )

    texto = ""
    try:
        async for chunk in client.stream_completion(
            tarea=tarea,
            system_prompt=prompt.contenido,
            user_content=user_content,
            user_id=USER_ID_BANCO,
            caso_id=caso.id,
            prompt_version=prompt.version,
            prompt_hash=prompt.hash,
        ):
            texto += chunk
    except JuliXError as exc:
        logger.error("Vridik/JuliX: fallo generando respuesta para %s: %s", caso.id, exc)
        texto = exc.partial_text or f"[JULIX_ERROR:{exc.status}] {exc}"
    return texto, tarea


async def calificar_con_juez(client: JuliXClient, caso: CasoEval, respuesta_julix: str) -> dict:
    """Llama al 'Claude juez' (misma infraestructura de client.py, tarea
    'evaluacion_juez') y valida que la salida sea el JSON esperado."""
    user_content = (
        f"Pregunta:\n{caso.pregunta}\n\n"
        f"Norma clave (única fuente permitida):\n{caso.norma_clave}\n\n"
        f"Patrón oro (respuesta esperada del despacho):\n{caso.respuesta_esperada}\n\n"
        f"Respuesta de JuliX a calificar:\n{respuesta_julix}"
    )

    texto = ""
    try:
        async for chunk in client.stream_completion(
            tarea="evaluacion_juez",
            system_prompt=JUEZ_SYSTEM_PROMPT,
            user_content=user_content,
            user_id=USER_ID_BANCO,
            caso_id=f"juez-{caso.id}",
            prompt_version=1,
            prompt_hash=_hash_prompt(JUEZ_SYSTEM_PROMPT),
        ):
            texto += chunk
    except JuliXError as exc:
        logger.error("Vridik/JuliX: fallo en el juez para %s: %s", caso.id, exc)
        # Fallo del juez nunca se traduce en aprobación silenciosa: score 0
        return {
            "score": 0, "precision_normativa": 0, "cita_correcta": False,
            "hallucination_flag": True, "comentario": f"Juez falló: {exc}",
        }

    try:
        return JuliXClient.validar_json(texto)
    except JuliXInvalidFormatError as exc:
        logger.error("Vridik/JuliX: salida del juez no es JSON válido para %s: %s", caso.id, exc)
        return {
            "score": 0, "precision_normativa": 0, "cita_correcta": False,
            "hallucination_flag": True, "comentario": "Salida del juez con formato inválido",
        }


async def registrar_resultado(db_connection, run_id: str, resultado: ResultadoCaso) -> None:
    query = """
        INSERT INTO julix_evals (
            caso_id, area, dificultad, model, score, precision_normativa,
            cita_correcta, hallucination_flag, comentario_juez, respuesta_julix,
            costo_usd_generacion, costo_usd_juez, run_id, created_at
        ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14)
    """
    await db_connection.execute(
        query,
        resultado.caso_id, resultado.area, resultado.dificultad, resultado.model,
        resultado.score, resultado.precision_normativa, resultado.cita_correcta,
        resultado.hallucination_flag, resultado.comentario_juez, resultado.respuesta_julix,
        resultado.costo_usd_generacion, resultado.costo_usd_juez, run_id,
        datetime.now(timezone.utc),
    )


async def _costo_por_caso_id(db_connection, caso_id: str) -> float | None:
    """Sustituye a julix.ledger.obtener_ultima_llamada() acá -- esa función
    filtra por user_id (UUID real de `users`), pero el banco de evaluación
    corre con USER_ID_BANCO=None (no hay usuario real asociado). caso_id
    identifica la llamada sin ambigüedad en esta corrida (único por caso, o
    'juez-{id}' para el juez), y desde el fix del 15-jul-2026 es TEXT, no
    UUID, así que acepta los IDs del banco tal cual ('UGPP-01', etc.)."""
    fila = await db_connection.fetchrow(
        "SELECT costo_usd FROM julix_calls WHERE caso_id = $1 ORDER BY created_at DESC LIMIT 1",
        caso_id,
    )
    return float(fila["costo_usd"]) if fila and fila["costo_usd"] is not None else None


def contrastar_flag_con_norma_clave(
    respuesta_julix: str, norma_clave: str, calificacion: dict,
) -> dict:
    """Regresión UGPP-07 (corrida s5 del 16-jul-2026): el juez marcó
    hallucination_flag=true por el Decreto 379/2026, que SÍ estaba en la
    norma_clave entregada — un falso positivo del juez, no una alucinación
    de JuliX. Este contraste es el chequeo MECÁNICO de ese patrón (mismo
    principio que julix/service.py::validar_citas_post_generacion: regex
    determinístico, no otra instrucción que el modelo pueda ignorar):

    Si el juez marcó alucinación pero TODAS las citas detectables en la
    respuesta de JuliX están respaldadas por la norma_clave, el flag es
    cuestionable. NUNCA se voltea en silencio (el regex es más burdo que el
    juez y la respuesta puede alucinar cifras/plazos que el regex no ve) —
    se anota `flag_cuestionado=True` y se deja rastro en el comentario para
    revisión humana, igual que la reclasificación manual del 16-jul.

    Reusa julix.service._claves_citables (claves normalizadas tipo
    'articulo:33' / 'decreto:379:2026') — una cita se considera respaldada
    si su clave aparece en la norma_clave, sin importar cómo se escribió.
    """
    calificacion = dict(calificacion)
    calificacion.setdefault("flag_cuestionado", False)
    if not calificacion.get("hallucination_flag"):
        return calificacion

    citadas = _claves_citables(respuesta_julix)
    if not citadas:
        # Sin citas detectables no hay nada que contrastar mecánicamente:
        # el flag del juez queda como está (puede ser por cifras/plazos).
        return calificacion

    respaldadas = set(_claves_citables(norma_clave).keys())
    sin_respaldo = {c: t for c, t in citadas.items() if c not in respaldadas}
    if sin_respaldo:
        # Hay al menos una cita que el regex tampoco encuentra en la
        # norma_clave: el flag del juez es plausible, no se cuestiona.
        return calificacion

    detalle = ", ".join(sorted(citadas.values()))
    calificacion["flag_cuestionado"] = True
    calificacion["comentario"] = (
        f"{calificacion.get('comentario', '')} "
        f"[flag_cuestionado] El juez marcó alucinación, pero todas las citas "
        f"detectables ({detalle}) están respaldadas por la norma_clave — "
        f"posible falso positivo del juez (patrón UGPP-07, 16-jul-2026). "
        f"Revisar manualmente antes de aceptar este veredicto."
    ).strip()
    logger.warning(
        "Vridik/JuliX eval: flag de alucinación cuestionado mecánicamente "
        "(todas las citas respaldadas por norma_clave) — revisar manualmente."
    )
    return calificacion


async def evaluar_caso(client: JuliXClient, db_connection, caso: CasoEval) -> ResultadoCaso:
    respuesta_julix, tarea = await generar_respuesta_julix(client, caso)
    costo_generacion = None
    if db_connection is not None:
        costo_generacion = await _costo_por_caso_id(db_connection, caso.id)

    calificacion = await calificar_con_juez(client, caso, respuesta_julix)
    calificacion = contrastar_flag_con_norma_clave(
        respuesta_julix, caso.norma_clave, calificacion,
    )
    costo_juez = None
    if db_connection is not None:
        costo_juez = await _costo_por_caso_id(db_connection, f"juez-{caso.id}")

    return ResultadoCaso(
        caso_id=caso.id,
        area=caso.area,
        dificultad=caso.dificultad,
        model=client.model_for(tarea),
        score=int(calificacion.get("score", 0)),
        precision_normativa=int(calificacion.get("precision_normativa", 0)),
        cita_correcta=bool(calificacion.get("cita_correcta", False)),
        hallucination_flag=bool(calificacion.get("hallucination_flag", False)),
        comentario_juez=str(calificacion.get("comentario", "")),
        respuesta_julix=respuesta_julix,
        costo_usd_generacion=costo_generacion,
        costo_usd_juez=costo_juez,
    )


async def correr_banco(
    excel_path: Path,
    *,
    client: JuliXClient | None,
    db_connection,
    commit: bool,
) -> ResumenBanco:
    todos_los_casos = leer_banco(excel_path)
    con_patron_oro = [c for c in todos_los_casos if c.respuesta_esperada]
    pendientes = len(todos_los_casos) - len(con_patron_oro)

    run_id = f"s5-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}-{uuid.uuid4().hex[:8]}"

    print(f"\n=== Vridik/JuliX — Banco de evaluación (corrida {run_id}) ===")
    print(f"Casos totales: {len(todos_los_casos)} | con patrón oro: {len(con_patron_oro)} | pendientes: {pendientes}")

    if pendientes:
        print(
            f"ATENCIÓN: {pendientes} caso(s) sin 'respuesta_esperada' — se OMITEN de esta corrida. "
            "Ver eval/guia_abogada.md para que Ana Luisa los complete."
        )

    if not commit:
        print("Modo dry-run: no se llama a Claude ni se escribe en julix_evals. Usa --commit para la corrida real.")
        return ResumenBanco(
            run_id=run_id, total_casos=len(todos_los_casos), casos_evaluados=0,
            casos_pendientes_patron_oro=pendientes, casos_aprobados=0, casos_con_alucinacion=0,
            porcentaje_aprobacion=0.0, gate_aprobado=False, resultados=[],
        )

    if client is None:
        raise RuntimeError("--commit requiere un JuliXClient inicializado (ver main())")

    resultados: list[ResultadoCaso] = []
    for caso in con_patron_oro:
        resultado = await evaluar_caso(client, db_connection, caso)
        resultados.append(resultado)
        if db_connection is not None:
            await registrar_resultado(db_connection, run_id, resultado)
        print(
            f"  [{resultado.caso_id:8s}] score={resultado.score} "
            f"alucinacion={resultado.hallucination_flag} cita_correcta={resultado.cita_correcta}"
        )

    aprobados = [r for r in resultados if r.score >= UMBRAL_APROBACION_CASO_SCORE and not r.hallucination_flag]
    con_alucinacion = [r for r in resultados if r.hallucination_flag]
    porcentaje = round(100 * len(aprobados) / len(resultados), 1) if resultados else 0.0
    gate_aprobado = (len(aprobados) / len(resultados) >= GATE_FASE1_PORCENTAJE) if resultados else False

    resumen = ResumenBanco(
        run_id=run_id,
        total_casos=len(todos_los_casos),
        casos_evaluados=len(resultados),
        casos_pendientes_patron_oro=pendientes,
        casos_aprobados=len(aprobados),
        casos_con_alucinacion=len(con_alucinacion),
        porcentaje_aprobacion=porcentaje,
        gate_aprobado=gate_aprobado,
        resultados=resultados,
    )

    print(f"\nResultado: {len(aprobados)}/{len(resultados)} aprobados ({porcentaje}%)")
    print(f"Casos con alucinación detectada: {len(con_alucinacion)}")
    print(
        f"GATE Fase 1 (>= {int(GATE_FASE1_PORCENTAJE * 100)}%): "
        + ("APROBADO" if gate_aprobado else "NO APROBADO -- ver S6, iteración de prompts")
    )
    return resumen


def main() -> int:
    parser = argparse.ArgumentParser(description="Vridik/JuliX — banco de evaluación (Gate de Fase 1, S5)")
    parser.add_argument("--excel", default=str(Path(__file__).parent / "banco_casos_vridik.xlsx"))
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Explícito (equivale al comportamiento por defecto): valida el banco, no llama a Claude ni escribe en BD",
    )
    parser.add_argument("--commit", action="store_true", help="Corrida real: llama a Claude y escribe en julix_evals")
    parser.add_argument("--environment", default="staging", choices=["staging", "production"])
    args = parser.parse_args()

    if args.dry_run and args.commit:
        print("ERROR: --dry-run y --commit son mutuamente excluyentes", file=sys.stderr)
        return 1

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: no se encontró el banco de casos en {excel_path}", file=sys.stderr)
        return 1

    client = None
    db_connection = None
    if args.commit:
        try:
            import asyncpg  # type: ignore
        except ImportError:
            print("ERROR: falta 'asyncpg' para --commit (pip install asyncpg)", file=sys.stderr)
            return 1
        import os

        database_url = os.environ.get("DATABASE_URL")
        if not database_url:
            print("ERROR: DATABASE_URL no configurado; requerido para --commit", file=sys.stderr)
            return 1

        async def _run():
            nonlocal client, db_connection
            db_connection = await asyncpg.connect(database_url)
            # Este script corre como "sistema" -- no está atado a ningún
            # despacho real (evalúa el banco completo, no la actividad de
            # un tenant), así que julix_calls.despacho_id queda siempre
            # NULL para sus registros. Sin este bypass explícito, la
            # política RLS de tenant isolation (WITH CHECK despacho_id::
            # text = app.despacho_id) rechaza cada INSERT del ledger --
            # bug real encontrado el 21-jul corriendo T3 contra producción
            # real (mismo síntoma que el de api/case_documents_endpoint.py,
            # causa distinta: ahí faltaba pasar un despacho_id real que sí
            # existía, acá no hay ninguno que pasar porque el caller
            # genuinamente no pertenece a un tenant).
            await db_connection.execute("SELECT set_config('app.bypass_rls', 'true', false)")
            client = JuliXClient(environment=args.environment, db_connection=db_connection)
            try:
                resumen = await correr_banco(excel_path, client=client, db_connection=db_connection, commit=True)
                return 0 if resumen.gate_aprobado or resumen.casos_evaluados == 0 else 1
            finally:
                await db_connection.close()

        return asyncio.run(_run())

    resumen = asyncio.run(correr_banco(excel_path, client=None, db_connection=None, commit=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
